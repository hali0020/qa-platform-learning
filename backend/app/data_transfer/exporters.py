from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO, StringIO
from typing import Iterable, Mapping
from uuid import UUID

from app.data_transfer.models import SpreadsheetDependencyError
from app.data_transfer.security import neutralize_spreadsheet_text, safe_download_name
from app.domain.models import Defect, TestCase
from app.schemas.data_transfer import TransferFormat


@dataclass(frozen=True, slots=True)
class TransferArtifact:
    filename: str
    media_type: str
    content: bytes
    row_count: int = 0


def export_test_cases(
    items: Iterable[TestCase],
    transfer_format: TransferFormat,
    *,
    suite_paths: Mapping[UUID, str] | None = None,
    generated_at: datetime | None = None,
) -> TransferArtifact:
    cases = sorted(items, key=lambda item: (item.created_at, str(item.id)))
    paths = suite_paths or {}
    main_rows = [
        {
            "id": str(item.id),
            "project_id": str(item.project_id),
            "suite_id": str(item.suite_id) if item.suite_id else "",
            "suite_path": paths.get(item.suite_id, "") if item.suite_id else "",
            "title": item.title,
            "preconditions": item.preconditions,
            "priority": item.priority.value,
            "case_type": item.case_type.value,
            "status": item.status.value,
            "tags": ";".join(item.tags),
            "steps_json": json.dumps(
                [step.model_dump() for step in item.steps],
                ensure_ascii=False,
                separators=(",", ":"),
            ),
            "created_at": item.created_at.isoformat(),
            "updated_at": item.updated_at.isoformat(),
        }
        for item in cases
    ]
    child_rows = [
        {
            "case_id": str(item.id),
            "position": position,
            "action": step.action,
            "expected_result": step.expected_result,
        }
        for item in cases
        for position, step in enumerate(item.steps, start=1)
    ]
    return _export(
        prefix="test-cases",
        transfer_format=transfer_format,
        main_sheet="test_cases",
        main_headers=(
            "id",
            "project_id",
            "suite_id",
            "suite_path",
            "title",
            "preconditions",
            "priority",
            "case_type",
            "status",
            "tags",
            "steps_json",
            "created_at",
            "updated_at",
        ),
        main_rows=main_rows,
        child_sheet="steps",
        child_headers=("case_id", "position", "action", "expected_result"),
        child_rows=child_rows,
        generated_at=generated_at,
    )


def export_defects(
    items: Iterable[Defect],
    transfer_format: TransferFormat,
    *,
    generated_at: datetime | None = None,
) -> TransferArtifact:
    defects = sorted(items, key=lambda item: (item.created_at, str(item.id)))
    main_rows = [
        {
            "id": str(item.id),
            "project_id": str(item.project_id),
            "case_id": str(item.case_id) if item.case_id else "",
            "execution_id": str(item.execution_id) if item.execution_id else "",
            "title": item.title,
            "description": item.description,
            "severity": item.severity.value,
            "priority": item.priority.value,
            "status": item.status.value,
            "reporter": item.reporter,
            "assignee": item.assignee,
            "environment": item.environment,
            "expected_result": item.expected_result,
            "actual_result": item.actual_result,
            "resolution": item.resolution,
            "reproduction_steps_json": json.dumps(
                item.reproduction_steps,
                ensure_ascii=False,
                separators=(",", ":"),
            ),
            "created_at": item.created_at.isoformat(),
            "updated_at": item.updated_at.isoformat(),
            "resolved_at": item.resolved_at.isoformat() if item.resolved_at else "",
            "closed_at": item.closed_at.isoformat() if item.closed_at else "",
        }
        for item in defects
    ]
    child_rows = [
        {"defect_id": str(item.id), "position": position, "step": step}
        for item in defects
        for position, step in enumerate(item.reproduction_steps, start=1)
    ]
    return _export(
        prefix="defects",
        transfer_format=transfer_format,
        main_sheet="defects",
        main_headers=(
            "id",
            "project_id",
            "case_id",
            "execution_id",
            "title",
            "description",
            "severity",
            "priority",
            "status",
            "reporter",
            "assignee",
            "environment",
            "expected_result",
            "actual_result",
            "resolution",
            "reproduction_steps_json",
            "created_at",
            "updated_at",
            "resolved_at",
            "closed_at",
        ),
        main_rows=main_rows,
        child_sheet="reproduction_steps",
        child_headers=("defect_id", "position", "step"),
        child_rows=child_rows,
        generated_at=generated_at,
    )


def _export(
    *,
    prefix: str,
    transfer_format: TransferFormat,
    main_sheet: str,
    main_headers: tuple[str, ...],
    main_rows: list[dict[str, object]],
    child_sheet: str,
    child_headers: tuple[str, ...],
    child_rows: list[dict[str, object]],
    generated_at: datetime | None,
) -> TransferArtifact:
    timestamp = (generated_at or datetime.now(timezone.utc)).astimezone(timezone.utc)
    stem = safe_download_name(f"{prefix}-{timestamp:%Y%m%d-%H%M%S}Z")
    if transfer_format == TransferFormat.CSV:
        content = _csv_bytes(main_headers, main_rows)
        return TransferArtifact(
            filename=f"{stem}.csv",
            media_type="text/csv; charset=utf-8",
            content=content,
            row_count=len(main_rows),
        )
    content = _xlsx_bytes(
        main_sheet,
        main_headers,
        main_rows,
        child_sheet,
        child_headers,
        child_rows,
    )
    return TransferArtifact(
        filename=f"{stem}.xlsx",
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        content=content,
        row_count=len(main_rows),
    )


def _csv_bytes(
    headers: tuple[str, ...],
    rows: list[dict[str, object]],
) -> bytes:
    stream = StringIO(newline="")
    stream.write("\ufeff")
    writer = csv.DictWriter(stream, fieldnames=headers, lineterminator="\r\n")
    writer.writeheader()
    for row in rows:
        writer.writerow(
            {
                key: neutralize_spreadsheet_text(value)
                for key, value in row.items()
            }
        )
    return stream.getvalue().encode("utf-8")


def _xlsx_bytes(
    main_sheet: str,
    main_headers: tuple[str, ...],
    main_rows: list[dict[str, object]],
    child_sheet: str,
    child_headers: tuple[str, ...],
    child_rows: list[dict[str, object]],
) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - exercised before dependency install
        raise SpreadsheetDependencyError() from exc

    workbook = Workbook(write_only=True)
    main = workbook.create_sheet(main_sheet)
    main.append(main_headers)
    for row in main_rows:
        main.append(
            [
                neutralize_spreadsheet_text(row.get(header, ""))
                for header in main_headers
            ]
        )
    child = workbook.create_sheet(child_sheet)
    child.append(child_headers)
    for row in child_rows:
        child.append(
            [
                neutralize_spreadsheet_text(row.get(header, ""))
                for header in child_headers
            ]
        )
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()
