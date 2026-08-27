from __future__ import annotations

import csv
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Any, Iterable

from app.data_transfer.models import (
    DataTransferFileError,
    ParsedImport,
    RawImportRow,
    SpreadsheetDependencyError,
)
from app.data_transfer.security import (
    MAX_PRIMARY_ROWS,
    MAX_TOTAL_ROWS,
    MAX_XLSX_COLUMNS,
    MAX_XLSX_METADATA_ROWS,
    sha256_hex,
    validate_upload,
)
from app.schemas.data_transfer import TransferEntity, TransferFormat

TEMPLATE_VERSION = "1"

MAIN_SHEETS = {
    TransferEntity.TEST_CASES: "test_cases",
    TransferEntity.DEFECTS: "defects",
}
CHILD_SHEETS = {
    TransferEntity.TEST_CASES: "steps",
    TransferEntity.DEFECTS: "reproduction_steps",
}


def parse_import_file(
    *,
    entity: TransferEntity,
    filename: str,
    content: bytes,
) -> ParsedImport:
    transfer_format = validate_upload(filename, content)
    digest = sha256_hex(content)
    if transfer_format == TransferFormat.CSV:
        return _parse_csv(entity, filename, content, digest)
    return _parse_xlsx(entity, filename, content, digest)


def _parse_csv(
    entity: TransferEntity,
    filename: str,
    content: bytes,
    digest: str,
) -> ParsedImport:
    try:
        text = content.decode("utf-8-sig", errors="strict")
    except UnicodeDecodeError as exc:
        raise DataTransferFileError(
            "invalid_csv_encoding",
            "CSV 必须使用 UTF-8 编码（允许 BOM）",
        ) from exc

    reader = csv.DictReader(StringIO(text, newline=""))
    headers = tuple(reader.fieldnames or ())
    _validate_headers(headers, MAIN_SHEETS[entity])
    rows: list[RawImportRow] = []
    versions: set[str] = set()
    for row_number, values in enumerate(reader, start=2):
        if None in values:
            raise DataTransferFileError(
                "csv_extra_columns",
                f"CSV 第 {row_number} 行包含表头之外的列",
            )
        normalized = {
            str(key).strip(): _normalize_cell(value)
            for key, value in values.items()
            if key is not None
        }
        if not any(value != "" for value in normalized.values()):
            continue
        versions.add(str(normalized.get("template_version", "")).strip())
        rows.append(
            RawImportRow(
                sheet=MAIN_SHEETS[entity],
                row=row_number,
                values=normalized,
            )
        )
        if len(rows) > MAX_PRIMARY_ROWS:
            raise DataTransferFileError(
                "too_many_rows",
                f"主数据最多允许 {MAX_PRIMARY_ROWS} 行",
            )
    if not rows:
        versions = {TEMPLATE_VERSION}
    if versions != {TEMPLATE_VERSION}:
        raise DataTransferFileError(
            "unsupported_template_version",
            f"CSV template_version 必须全部为 {TEMPLATE_VERSION}",
        )
    return ParsedImport(
        entity=entity,
        filename=filename,
        transfer_format=TransferFormat.CSV,
        sha256=digest,
        template_version=TEMPLATE_VERSION,
        main_sheet=MAIN_SHEETS[entity],
        main_headers=headers,
        main_rows=tuple(rows),
    )


def _parse_xlsx(
    entity: TransferEntity,
    filename: str,
    content: bytes,
    digest: str,
) -> ParsedImport:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised before dependency install
        raise SpreadsheetDependencyError() from exc

    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise DataTransferFileError(
            "invalid_xlsx_workbook",
            "无法读取 XLSX 工作簿",
        ) from exc
    try:
        required = {"_meta", MAIN_SHEETS[entity], CHILD_SHEETS[entity]}
        missing = sorted(required - set(workbook.sheetnames))
        if missing:
            raise DataTransferFileError(
                "missing_xlsx_sheet",
                "XLSX 缺少工作表: " + ", ".join(missing),
            )
        allowed = required | {"_lists", "说明"}
        unexpected = sorted(set(workbook.sheetnames) - allowed)
        if unexpected:
            raise DataTransferFileError(
                "unexpected_xlsx_sheet",
                "XLSX 包含未知工作表: " + ", ".join(unexpected),
            )

        metadata = _read_metadata(workbook["_meta"])
        if metadata.get("entity") != entity.value:
            raise DataTransferFileError(
                "template_entity_mismatch",
                "模板实体类型与导入入口不匹配",
            )
        version = metadata.get("template_version", "")
        if version != TEMPLATE_VERSION:
            raise DataTransferFileError(
                "unsupported_template_version",
                f"XLSX template_version 必须为 {TEMPLATE_VERSION}",
            )

        main_headers, main_rows = _read_sheet(
            workbook[MAIN_SHEETS[entity]],
            max_data_rows=MAX_PRIMARY_ROWS,
        )
        child_headers, child_rows = _read_sheet(
            workbook[CHILD_SHEETS[entity]],
            max_data_rows=MAX_TOTAL_ROWS,
        )
        if len(main_rows) > MAX_PRIMARY_ROWS:
            raise DataTransferFileError(
                "too_many_rows",
                f"主数据最多允许 {MAX_PRIMARY_ROWS} 行",
            )
        if len(main_rows) + len(child_rows) > MAX_TOTAL_ROWS:
            raise DataTransferFileError(
                "too_many_total_rows",
                f"工作簿总数据最多允许 {MAX_TOTAL_ROWS} 行",
            )
        return ParsedImport(
            entity=entity,
            filename=filename,
            transfer_format=TransferFormat.XLSX,
            sha256=digest,
            template_version=version,
            main_sheet=MAIN_SHEETS[entity],
            main_headers=main_headers,
            main_rows=main_rows,
            child_sheet=CHILD_SHEETS[entity],
            child_headers=child_headers,
            child_rows=child_rows,
            metadata=metadata,
        )
    finally:
        workbook.close()


def _read_metadata(sheet) -> dict[str, str]:
    _validate_sheet_dimensions(
        sheet,
        max_rows=MAX_XLSX_METADATA_ROWS,
        max_columns=16,
    )
    metadata: dict[str, str] = {}
    for index, cells in enumerate(sheet.iter_rows(values_only=False), start=1):
        if index > MAX_XLSX_METADATA_ROWS:
            raise DataTransferFileError(
                "xlsx_metadata_too_large",
                f"工作表 {sheet.title} 最多允许 {MAX_XLSX_METADATA_ROWS} 行",
            )
        _reject_formulas(cells, sheet.title, index)
        values = [_normalize_cell(cell.value) for cell in cells]
        if index == 1 and values[:2] == ["key", "value"]:
            continue
        if len(values) >= 2 and values[0]:
            metadata[values[0]] = values[1]
    return metadata


def _read_sheet(
    sheet,
    *,
    max_data_rows: int,
) -> tuple[tuple[str, ...], tuple[RawImportRow, ...]]:
    _validate_sheet_dimensions(
        sheet,
        max_rows=max_data_rows + 1,
        max_columns=MAX_XLSX_COLUMNS,
    )
    iterator = sheet.iter_rows(values_only=False)
    try:
        header_cells = next(iterator)
    except StopIteration as exc:
        raise DataTransferFileError(
            "missing_header",
            f"工作表 {sheet.title} 缺少表头",
        ) from exc
    _reject_formulas(header_cells, sheet.title, 1)
    headers = tuple(_normalize_cell(cell.value).strip() for cell in header_cells)
    _validate_headers(headers, sheet.title)
    rows: list[RawImportRow] = []
    for row_number, cells in enumerate(iterator, start=2):
        if row_number > max_data_rows + 1:
            raise DataTransferFileError(
                "xlsx_logical_rows_too_large",
                f"工作表 {sheet.title} 逻辑行数超过允许上限",
            )
        _reject_formulas(cells, sheet.title, row_number)
        values = [_normalize_cell(cell.value) for cell in cells]
        if not any(value != "" for value in values):
            continue
        if len(values) > len(headers) and any(
            value != "" for value in values[len(headers) :]
        ):
            raise DataTransferFileError(
                "xlsx_extra_columns",
                f"工作表 {sheet.title} 第 {row_number} 行包含表头之外的列",
            )
        mapped = {
            header: values[index] if index < len(values) else ""
            for index, header in enumerate(headers)
        }
        rows.append(
            RawImportRow(
                sheet=sheet.title,
                row=row_number,
                values=mapped,
            )
        )
    return headers, tuple(rows)


def _validate_sheet_dimensions(
    sheet,
    *,
    max_rows: int,
    max_columns: int,
) -> None:
    logical_rows = int(sheet.max_row or 0)
    logical_columns = int(sheet.max_column or 0)
    if logical_rows > max_rows:
        raise DataTransferFileError(
            "xlsx_logical_rows_too_large",
            f"工作表 {sheet.title} 逻辑行数不能超过 {max_rows}",
        )
    if logical_columns > max_columns:
        raise DataTransferFileError(
            "xlsx_logical_columns_too_large",
            f"工作表 {sheet.title} 逻辑列数不能超过 {max_columns}",
        )


def _reject_formulas(cells: Iterable[Any], sheet: str, row: int) -> None:
    for cell in cells:
        if getattr(cell, "data_type", None) == "f":
            raise DataTransferFileError(
                "formula_not_allowed",
                f"工作表 {sheet} 第 {row} 行包含公式；导入文件只允许静态值",
            )


def _validate_headers(headers: tuple[str, ...], sheet: str) -> None:
    if not headers:
        raise DataTransferFileError("missing_header", f"{sheet} 缺少表头")
    if any(not header for header in headers):
        raise DataTransferFileError("blank_header", f"{sheet} 包含空表头")
    if len(set(headers)) != len(headers):
        raise DataTransferFileError("duplicate_header", f"{sheet} 包含重复表头")


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)
