from __future__ import annotations

import csv
from io import BytesIO, StringIO

from app.data_transfer.models import SpreadsheetDependencyError
from app.data_transfer.parsers import TEMPLATE_VERSION
from app.data_transfer.security import safe_download_name
from app.schemas.data_transfer import TransferEntity, TransferFormat

CSV_HEADERS = {
    TransferEntity.TEST_CASES: (
        "template_version",
        "row_key",
        "suite_id",
        "suite_path",
        "title",
        "preconditions",
        "priority",
        "case_type",
        "tags",
        "steps_json",
    ),
    TransferEntity.DEFECTS: (
        "template_version",
        "row_key",
        "title",
        "description",
        "severity",
        "priority",
        "reporter",
        "assignee",
        "environment",
        "case_id",
        "execution_id",
        "expected_result",
        "actual_result",
        "reproduction_steps_json",
    ),
}

XLSX_MAIN_HEADERS = {
    TransferEntity.TEST_CASES: (
        "row_key",
        "suite_id",
        "suite_path",
        "title",
        "preconditions",
        "priority",
        "case_type",
        "tags",
    ),
    TransferEntity.DEFECTS: (
        "row_key",
        "title",
        "description",
        "severity",
        "priority",
        "reporter",
        "assignee",
        "environment",
        "case_id",
        "execution_id",
        "expected_result",
        "actual_result",
    ),
}

XLSX_CHILD_HEADERS = {
    TransferEntity.TEST_CASES: ("row_key", "position", "action", "expected_result"),
    TransferEntity.DEFECTS: ("row_key", "position", "step"),
}

MAIN_SHEET_NAMES = {
    TransferEntity.TEST_CASES: "test_cases",
    TransferEntity.DEFECTS: "defects",
}
CHILD_SHEET_NAMES = {
    TransferEntity.TEST_CASES: "steps",
    TransferEntity.DEFECTS: "reproduction_steps",
}


def build_import_template(
    entity: TransferEntity,
    transfer_format: TransferFormat,
):
    """Return a small immutable artifact without touching disk."""

    from app.data_transfer.exporters import TransferArtifact

    stem = safe_download_name(f"{entity.value}-import-template-v{TEMPLATE_VERSION}")
    if transfer_format == TransferFormat.CSV:
        stream = StringIO(newline="")
        stream.write("\ufeff")
        csv.writer(stream, lineterminator="\r\n").writerow(CSV_HEADERS[entity])
        return TransferArtifact(
            filename=f"{stem}.csv",
            media_type="text/csv; charset=utf-8",
            content=stream.getvalue().encode("utf-8"),
        )

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:  # pragma: no cover - exercised before dependency install
        raise SpreadsheetDependencyError() from exc

    workbook = Workbook(write_only=False)
    default = workbook.active
    workbook.remove(default)
    meta = workbook.create_sheet("_meta")
    meta.append(("key", "value"))
    meta.append(("entity", entity.value))
    meta.append(("template_version", TEMPLATE_VERSION))
    meta.sheet_state = "hidden"

    main = workbook.create_sheet(MAIN_SHEET_NAMES[entity])
    main.append(XLSX_MAIN_HEADERS[entity])
    child = workbook.create_sheet(CHILD_SHEET_NAMES[entity])
    child.append(XLSX_CHILD_HEADERS[entity])
    notes = workbook.create_sheet("说明")
    notes.append(("规则", "说明"))
    notes.append(("row_key", "文件内唯一；用于主表和步骤表关联，不写入业务实体"))
    notes.append(("状态", "导入用例统一创建为 draft；导入缺陷统一创建为 open"))
    notes.append(("套件", "suite_id 与 suite_path 可二选一；同时填写时必须指向同一套件"))
    notes.append(("公式", "为安全起见，任何公式单元格都会拒绝导入"))
    for sheet in (main, child, notes):
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="355070")
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(
                42,
                max(12, max(len(str(cell.value or "")) for cell in column) + 2),
            )
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return TransferArtifact(
        filename=f"{stem}.xlsx",
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        content=buffer.getvalue(),
    )
